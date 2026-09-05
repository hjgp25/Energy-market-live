#!/usr/bin/env python3
"""
Energy Market Live — official Baker Hughes updater.

Primary source:
  Baker Hughes North America Rig Count weekly Excel report.

It automatically:
- finds the newest Baker Hughes North America report
- downloads the Excel workbook
- finds the breakdown table
- uses the latest and previous publish dates
- calculates U.S./Canada totals and week-over-week changes
- builds state, basin, drill-for, trajectory and location breakdowns
- preserves the last good values if Baker Hughes is temporarily unavailable

Expected output: rigcount.json in the repository root.
"""

from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from dateutil import parser as dateparser
from openpyxl import load_workbook

OUT = Path("rigcount.json")

REPORT_PAGES = [
    "https://rigcount.bakerhughes.com/na-rig-count/",
    "https://bakerhughesrigcount.gcs-web.com/na-rig-count/",
]

SUMMARY_PAGES = [
    "https://rigcount.bakerhughes.com/",
    "https://bakerhughesrigcount.gcs-web.com/",
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; EnergyMarketLive/2.0; +GitHubPages)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control": "no-cache",
}

STATE_CODES = {
    "Alabama":"AL","Alaska":"AK","Arizona":"AZ","Arkansas":"AR","California":"CA","Colorado":"CO",
    "Connecticut":"CT","Delaware":"DE","Florida":"FL","Georgia":"GA","Hawaii":"HI","Idaho":"ID",
    "Illinois":"IL","Indiana":"IN","Iowa":"IA","Kansas":"KS","Kentucky":"KY","Louisiana":"LA",
    "Maine":"ME","Maryland":"MD","Massachusetts":"MA","Michigan":"MI","Minnesota":"MN","Mississippi":"MS",
    "Missouri":"MO","Montana":"MT","Nebraska":"NE","Nevada":"NV","New Hampshire":"NH","New Jersey":"NJ",
    "New Mexico":"NM","New York":"NY","North Carolina":"NC","North Dakota":"ND","Ohio":"OH","Oklahoma":"OK",
    "Oregon":"OR","Pennsylvania":"PA","Rhode Island":"RI","South Carolina":"SC","South Dakota":"SD",
    "Tennessee":"TN","Texas":"TX","Utah":"UT","Vermont":"VT","Virginia":"VA","Washington":"WA",
    "West Virginia":"WV","Wisconsin":"WI","Wyoming":"WY"
}

ALIASES = {
    "country": {"country"},
    "state": {"state/province","state province","state","province"},
    "county": {"county"},
    "basin": {"basin"},
    "drill_for": {"drill for","drillfor","drill for type","drill_for"},
    "location": {"location"},
    "trajectory": {"trajectory"},
    "publish_date": {"us_publishdate","us publishdate","us publish date","publish date","date"},
    "rig_count": {"rig count value","rig_count_value","rig count","count"},
}

def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def norm(v) -> str:
    return clean(v).lower().replace("\n"," ").strip()

def as_int(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(round(v))
    m = re.search(r"[+-]?\d+(?:\.\d+)?", clean(v).replace(",",""))
    return int(round(float(m.group()))) if m else 0

def parse_date(v):
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    try:
        return dateparser.parse(clean(v), fuzzy=True).replace(tzinfo=None)
    except Exception:
        return None

def read_existing():
    if not OUT.exists():
        return {}
    try:
        return json.loads(OUT.read_text(encoding="utf-8"))
    except Exception:
        return {}

def get(url, **kwargs):
    r = requests.get(url, headers=HEADERS, timeout=45, **kwargs)
    r.raise_for_status()
    return r

def discover_latest_xlsx():
    errors = []
    for page in REPORT_PAGES:
        try:
            html = get(page).text
            soup = BeautifulSoup(html, "html.parser")

            # Prefer a link near "New Report".
            candidates = []
            for a in soup.find_all("a", href=True):
                href = a.get("href","")
                text = clean(a.get_text(" ", strip=True))
                if "/static-files/" in href:
                    score = 2 if "new report" in text.lower() else 1
                    candidates.append((score, href, text))

            if not candidates:
                # Regex fallback for pages where the anchor text is unusual.
                for href in re.findall(r'href=["\']([^"\']*/static-files/[^"\']+)["\']', html, re.I):
                    candidates.append((1, href, ""))

            if not candidates:
                raise RuntimeError("No Baker Hughes static-file report link found")

            candidates.sort(key=lambda x: x[0], reverse=True)
            return urljoin(page, candidates[0][1])
        except Exception as exc:
            errors.append(f"{page}: {exc}")

    raise RuntimeError("Could not discover Baker Hughes Excel report: " + " | ".join(errors))

def header_map(row):
    result = {}
    normalized = [norm(x) for x in row]
    for key, aliases in ALIASES.items():
        for i, value in enumerate(normalized):
            if value in aliases:
                result[key] = i
                break
    return result

def find_breakdown_sheet(wb):
    best = None
    best_score = -1

    for ws in wb.worksheets:
        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=True), start=1):
            h = header_map(row)
            score = len(h)
            # We must have these two to be useful.
            if "publish_date" in h and "rig_count" in h and score > best_score:
                best = (ws, row_num, h)
                best_score = score

    if not best or best_score < 5:
        raise RuntimeError("Could not locate the NAM breakdown table in the workbook")
    return best

def load_records(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws, header_row, hm = find_breakdown_sheet(wb)

    records = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        def cell(key):
            idx = hm.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        d = parse_date(cell("publish_date"))
        count = as_int(cell("rig_count"))
        country = clean(cell("country"))
        if d is None or not country:
            continue

        records.append({
            "date": d,
            "country": country,
            "state": clean(cell("state")),
            "county": clean(cell("county")),
            "basin": clean(cell("basin")),
            "drill_for": clean(cell("drill_for")),
            "location": clean(cell("location")),
            "trajectory": clean(cell("trajectory")),
            "count": count,
        })

    if not records:
        raise RuntimeError("Breakdown table contained no usable records")
    return records

def is_us(country):
    c = norm(country).replace(".","")
    return c in {"us","usa","united states","united states of america"}

def is_canada(country):
    return norm(country) == "canada"

def group(records, field, country_test):
    totals = defaultdict(int)
    for r in records:
        if country_test(r["country"]):
            label = clean(r.get(field))
            if label:
                totals[label] += r["count"]
    return totals

def total(records, country_test):
    return sum(r["count"] for r in records if country_test(r["country"]))

def make_breakdown(current, previous, field, country_test):
    cur = group(current, field, country_test)
    prev = group(previous, field, country_test)
    out = []
    for name, count in cur.items():
        out.append({
            "name": name,
            "count": count,
            "change": count - prev.get(name, 0)
        })
    return sorted(out, key=lambda x: x["count"], reverse=True)

def dict_breakdown(current, previous, field, country_test):
    cur = group(current, field, country_test)
    prev = group(previous, field, country_test)
    out = {}
    for name, count in cur.items():
        key = norm(name).replace(" ","_").replace("-","_").replace("/","_")
        out[key] = {
            "count": count,
            "change": count - prev.get(name, 0)
        }
    return out

def scrape_official_weekly(existing):
    url = discover_latest_xlsx()
    print("Official Baker Hughes report:", url)
    xlsx = get(url).content
    records = load_records(xlsx)

    dates = sorted({r["date"] for r in records})
    if len(dates) < 2:
        raise RuntimeError("Need at least two publish dates to calculate week-over-week changes")

    latest_date = dates[-1]
    previous_date = dates[-2]
    current = [r for r in records if r["date"] == latest_date]
    previous = [r for r in records if r["date"] == previous_date]

    us_now = total(current, is_us)
    us_prev = total(previous, is_us)
    ca_now = total(current, is_canada)
    ca_prev = total(previous, is_canada)

    states_raw = make_breakdown(current, previous, "state", is_us)
    states = []
    for x in states_raw:
        name = x["name"]
        if name in STATE_CODES:
            states.append({
                "code": STATE_CODES[name],
                "name": name,
                "count": x["count"],
                "change": x["change"],
            })

    basins = make_breakdown(current, previous, "basin", is_us)
    drill_for = dict_breakdown(current, previous, "drill_for", is_us)
    trajectory = dict_breakdown(current, previous, "trajectory", is_us)
    location = dict_breakdown(current, previous, "location", is_us)

    report_date = latest_date.strftime("%B %-d, %Y")
    prev_date = previous_date.strftime("%B %-d, %Y")

    return {
        "source": "Baker Hughes Rig Count",
        "official_report_url": url,
        "report_date": report_date,
        "previous_report_date": prev_date,
        "state_report_date": report_date,
        "basin_report_date": report_date,
        "us": {"count": us_now, "change": us_now - us_prev},
        "canada": {"count": ca_now, "change": ca_now - ca_prev},
        "states": states,
        "basins": basins,
        "drill_for": drill_for,
        "trajectory": trajectory,
        "location": location,
    }

def scrape_summary(existing):
    for page in SUMMARY_PAGES:
        try:
            soup = BeautifulSoup(get(page).text, "html.parser")
            result = {}
            for table in soup.find_all("table"):
                for tr in table.find_all("tr"):
                    cells = [clean(x.get_text(" ", strip=True)) for x in tr.find_all(["th","td"])]
                    if len(cells) < 4:
                        continue
                    label = norm(cells[0]).replace(".","")
                    item = {
                        "last_count": cells[1],
                        "count": as_int(cells[2]),
                        "change": as_int(cells[3]),
                    }
                    if label in {"us","usa"} or label.startswith("us "):
                        result["us"] = item
                    elif label.startswith("canada"):
                        result["canada"] = item
                    elif label.startswith("international"):
                        result["international"] = item
            if result:
                return result
        except Exception as exc:
            print("Summary mirror warning:", page, exc)
    return {}

def main():
    existing = read_existing()
    merged = dict(existing)

    try:
        weekly = scrape_official_weekly(existing)
        merged.update(weekly)
        print("SUCCESS: official Baker Hughes weekly Excel parsed.")
    except Exception as exc:
        print("OFFICIAL WEEKLY WARNING:", repr(exc))
        print("Keeping last good weekly/state/basin data.")

    # International is monthly and comes from the Baker Hughes summary page.
    try:
        summary = scrape_summary(existing)
        if summary.get("international", {}).get("count") is not None:
            merged["international"] = summary["international"]

        # Only use summary U.S./Canada if weekly workbook was unavailable.
        if "us" not in merged and summary.get("us"):
            merged["us"] = summary["us"]
        if "canada" not in merged and summary.get("canada"):
            merged["canada"] = summary["canada"]
    except Exception as exc:
        print("SUMMARY WARNING:", repr(exc))

    merged["source"] = "Baker Hughes Rig Count"
    merged["fetched_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")

    OUT.write_text(json.dumps(merged, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(merged, indent=2))
    print("DONE: rigcount.json written.")

if __name__ == "__main__":
    main()
